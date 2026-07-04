"""Write the canonical Excel-pasteable Siemens wide TSV export."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from .config import DAYS_ROWS, OUTPUT_ROWS, PERCENT_ROWS, RATIO_ROWS
from .models import QuarterData, YearData
from .periods import quarter_sort_key, year_sort_key


def format_cell(value: int | float | str | None, row: str | None = None) -> str:
    """Format audit values into row-specific TSV cell text."""
    if value is None:
        return ""
    if isinstance(value, float):
        if row in RATIO_ROWS:
            return f"{value:.2f}"
        if row in DAYS_ROWS:
            return f"{value:.1f}"
        if row in PERCENT_ROWS or row is None:
            return f"{value:.1%}"
        return f"{value:.1f}"
    if isinstance(value, int):
        return f"{value:,}"
    return str(value)


def write_tsv(path: Path, quarters: dict[str, QuarterData]) -> None:
    """Write quarters in chronological order using the fixed output template."""
    columns = sorted(quarters, key=quarter_sort_key)
    lines = ["Financial Years\t" + "\t".join(columns)]
    for row in OUTPUT_ROWS:
        if row is None:
            lines.append("")
            continue
        cells = [row] + [format_cell(quarters[column].values.get(row), row) for column in columns]
        lines.append("\t".join(without_trailing_empty_cells(cells)))
    path.write_text("\n".join(lines) + "\n")


def write_workbook(path: Path, quarters: dict[str, QuarterData], years: dict[str, YearData]) -> None:
    """Write a two-sheet workbook with quarterly and derived yearly exports."""
    workbook = Workbook()
    quarterly_sheet = workbook.active
    quarterly_sheet.title = "Quarterly"
    _write_grid(quarterly_sheet, sorted(quarters, key=quarter_sort_key), quarters)
    yearly_sheet = workbook.create_sheet("Yearly")
    _write_grid(yearly_sheet, sorted(years, key=year_sort_key), years)
    workbook.save(path)


def _write_grid(sheet: object, columns: list[str], data: dict[str, QuarterData] | dict[str, YearData]) -> None:
    """Write one worksheet grid using the TSV row order and cell formatting."""
    sheet.cell(row=1, column=1, value="Financial Years")
    for column_index, column in enumerate(columns, start=2):
        sheet.cell(row=1, column=column_index, value=column)
    for row_index, output_row in enumerate(OUTPUT_ROWS, start=2):
        if output_row is None:
            continue
        sheet.cell(row=row_index, column=1, value=output_row)
        for column_index, column in enumerate(columns, start=2):
            sheet.cell(
                row=row_index,
                column=column_index,
                value=format_cell(data[column].values.get(output_row), output_row),
            )


def without_trailing_empty_cells(cells: list[str]) -> list[str]:
    """Remove trailing blanks so TSV rows do not end with tab characters."""
    trimmed = list(cells)
    while trimmed and trimmed[-1] == "":
        trimmed.pop()
    return trimmed
