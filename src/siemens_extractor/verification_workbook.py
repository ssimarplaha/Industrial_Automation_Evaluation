"""Verify the derived Siemens Excel workbook export."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .config import OUTPUT_ROWS
from .periods import year_sort_key
from .verification_rules import calculated_values_equal
from .writer import format_cell
from .yearly import calculate_years


def verify_workbook_outputs(
    workbook_path: Path,
    tsv_path: Path,
    audit: dict[str, Any],
) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    """Verify workbook structure, sheet values, and annual audit replay."""
    issues: list[dict[str, Any]] = []
    checked_values: list[dict[str, Any]] = []
    metadata = {
        "workbook_path": str(workbook_path),
        "workbook_checked": False,
        "year_columns_checked": 0,
        "yearly_exported_non_empty_cells_checked": 0,
    }
    if not workbook_path.exists():
        _add_issue(issues, "missing_workbook", f"Workbook is missing: {workbook_path}.")
        return metadata, checked_values, issues

    workbook = load_workbook(workbook_path, data_only=True)
    metadata["workbook_checked"] = True
    expected_sheets = ["Quarterly", "Yearly"]
    if workbook.sheetnames != expected_sheets:
        _add_issue(
            issues,
            "workbook_sheet_mismatch",
            "Workbook must contain exactly Quarterly and Yearly sheets.",
            expected_sheets=expected_sheets,
            actual_sheets=workbook.sheetnames,
        )
        return metadata, checked_values, issues

    _verify_quarterly_sheet(workbook["Quarterly"], tsv_path, checked_values, issues)
    yearly_metadata = _verify_yearly_sheet(workbook["Yearly"], audit, checked_values, issues)
    metadata.update(yearly_metadata)
    _verify_yearly_audit_replay(audit, checked_values, issues)
    return metadata, checked_values, issues


def _verify_quarterly_sheet(
    sheet: Any,
    tsv_path: Path,
    checked_values: list[dict[str, Any]],
    issues: list[dict[str, Any]],
) -> None:
    """Check that the Quarterly workbook sheet mirrors the TSV grid."""
    lines = tsv_path.read_text().splitlines()
    expected_width = len(lines[0].split("\t")) if lines else 0
    for row_index, line in enumerate(lines, start=1):
        cells = line.split("\t")
        cells += [""] * max(0, expected_width - len(cells))
        for column_index, expected in enumerate(cells[:expected_width], start=1):
            actual = _cell_text(sheet.cell(row=row_index, column=column_index).value)
            checked_values.append(
                {
                    "check": "workbook_quarterly_cell",
                    "row": row_index,
                    "column": column_index,
                    "workbook_value": actual,
                    "tsv_value": expected,
                    "passed": actual == expected,
                }
            )
            if actual != expected:
                _add_issue(
                    issues,
                    "workbook_quarterly_tsv_mismatch",
                    "Quarterly workbook cell does not match TSV.",
                    row_number=row_index,
                    column_number=column_index,
                    workbook_value=actual,
                    tsv_value=expected,
                )


def _verify_yearly_sheet(
    sheet: Any,
    audit: dict[str, Any],
    checked_values: list[dict[str, Any]],
    issues: list[dict[str, Any]],
) -> dict[str, Any]:
    """Check the Yearly workbook sheet against annual audit values."""
    year_columns = audit.get("metadata", {}).get("year_columns", [])
    actual_columns = [_cell_text(sheet.cell(row=1, column=index).value) for index in range(2, len(year_columns) + 2)]
    if _cell_text(sheet.cell(row=1, column=1).value) != "Financial Years":
        _add_issue(issues, "invalid_yearly_header", "Yearly sheet header must start with Financial Years.")
    if actual_columns != year_columns:
        _add_issue(
            issues,
            "yearly_column_mismatch",
            "Yearly sheet columns do not match audit metadata.",
            workbook_columns=actual_columns,
            audit_columns=year_columns,
        )

    years = audit.get("years", {})
    non_empty = 0
    for row_index, output_row in enumerate(OUTPUT_ROWS, start=2):
        if output_row is None:
            _verify_blank_separator(sheet, row_index, len(year_columns) + 1, issues)
            continue
        actual_row = _cell_text(sheet.cell(row=row_index, column=1).value)
        if actual_row != output_row:
            _add_issue(
                issues,
                "yearly_row_mismatch",
                "Yearly sheet row order does not match output template.",
                row_number=row_index,
                workbook_row=actual_row,
                expected_row=output_row,
            )
        for column_index, year_code in enumerate(year_columns, start=2):
            actual = _cell_text(sheet.cell(row=row_index, column=column_index).value)
            year = years.get(year_code, {})
            audit_value = year.get("values", {}).get(output_row)
            expected = format_cell(audit_value, output_row)
            non_empty += 1 if actual else 0
            checked_values.append(
                {
                    "check": "workbook_yearly_cell",
                    "year": year_code,
                    "row": output_row,
                    "workbook_value": actual,
                    "audit_value": audit_value,
                    "passed": actual == expected,
                }
            )
            if actual != expected:
                _add_issue(
                    issues,
                    "workbook_yearly_audit_mismatch",
                    f"{year_code} {output_row}: Yearly workbook value does not match audit.",
                    year=year_code,
                    row=output_row,
                    workbook_value=actual,
                    audit_value=audit_value,
                    expected_workbook_value=expected,
                )
            if actual and output_row not in year.get("sources", {}):
                _add_issue(
                    issues,
                    "missing_yearly_exported_source",
                    f"{year_code} {output_row}: non-empty Yearly workbook value has no audit source.",
                    year=year_code,
                    row=output_row,
                    workbook_value=actual,
                )
    return {
        "year_columns_checked": len(year_columns),
        "yearly_exported_non_empty_cells_checked": non_empty,
    }


def _verify_blank_separator(sheet: Any, row_index: int, width: int, issues: list[dict[str, Any]]) -> None:
    """Check that a separator row does not contain workbook values."""
    for column_index in range(1, width + 1):
        if _cell_text(sheet.cell(row=row_index, column=column_index).value):
            _add_issue(
                issues,
                "nonblank_yearly_separator_row",
                "Yearly blank separator row contains values.",
                row_number=row_index,
            )
            return


def _verify_yearly_audit_replay(
    audit: dict[str, Any],
    checked_values: list[dict[str, Any]],
    issues: list[dict[str, Any]],
) -> None:
    """Replay annual values from quarterly audit data and compare the stored audit."""
    expected_years = calculate_years(audit.get("quarters", {}))
    expected_columns = sorted(expected_years, key=year_sort_key)
    audit_columns = audit.get("metadata", {}).get("year_columns", [])
    if audit_columns != expected_columns:
        _add_issue(
            issues,
            "year_audit_column_mismatch",
            "Audit year columns do not match complete years derived from quarter data.",
            audit_columns=audit_columns,
            expected_columns=expected_columns,
        )

    audit_years = audit.get("years", {})
    for code in sorted(set(expected_columns) | set(audit_years), key=year_sort_key):
        expected = expected_years.get(code)
        actual = audit_years.get(code)
        if expected is None or actual is None:
            _add_issue(issues, "missing_year_audit", f"Audit is missing or has unexpected year {code}.", year=code)
            continue
        _verify_one_year_audit(code, expected.values, actual, checked_values, issues)


def _verify_one_year_audit(
    code: str,
    expected_values: dict[str, Any],
    actual: dict[str, Any],
    checked_values: list[dict[str, Any]],
    issues: list[dict[str, Any]],
) -> None:
    """Compare one annual audit column against replayed values and source records."""
    actual_values = actual.get("values", {})
    for row in sorted(set(expected_values) | set(actual_values)):
        expected = expected_values.get(row)
        audit_value = actual_values.get(row)
        source = actual.get("sources", {}).get(row)
        checked_values.append(
            {
                "check": "yearly_audit_replay",
                "year": code,
                "row": row,
                "audit_value": audit_value,
                "expected_value": expected,
                "passed": calculated_values_equal(expected, audit_value),
            }
        )
        if not calculated_values_equal(expected, audit_value):
            _add_issue(
                issues,
                "yearly_audit_mismatch",
                f"{code} {row}: annual audit value does not replay from quarterly audit data.",
                year=code,
                row=row,
                audit_value=audit_value,
                expected_value=expected,
            )
        if audit_value is None:
            continue
        if source is None:
            _add_issue(issues, "missing_yearly_audit_source", f"{code} {row}: annual audit value has no source.")
            continue
        _verify_year_source(code, row, audit_value, source, issues)


def _verify_year_source(
    code: str,
    row: str,
    value: Any,
    source: dict[str, Any],
    issues: list[dict[str, Any]],
) -> None:
    """Validate one annual calculated source record."""
    if source.get("source_type") != "calculated":
        _add_issue(
            issues,
            "invalid_yearly_source_type",
            f"{code} {row}: annual source type must be calculated.",
            year=code,
            row=row,
            source_type=source.get("source_type"),
        )
    if source.get("parser_family") != "yearly":
        _add_issue(
            issues,
            "invalid_yearly_parser_family",
            f"{code} {row}: annual source parser family must be yearly.",
            year=code,
            row=row,
            parser_family=source.get("parser_family"),
        )
    if source.get("normalized_row") != row:
        _add_issue(
            issues,
            "yearly_normalized_row_mismatch",
            f"{code} {row}: annual source normalized row does not match.",
            year=code,
            row=row,
            normalized_row=source.get("normalized_row"),
        )
    normalized_value = source.get("normalized_value")
    if normalized_value != value and normalized_value != format_cell(value, row):
        _add_issue(
            issues,
            "yearly_normalized_value_mismatch",
            f"{code} {row}: annual source normalized value does not match.",
            year=code,
            row=row,
            normalized_value=normalized_value,
            audit_value=value,
        )


def _cell_text(value: Any) -> str:
    """Normalize workbook cell values to comparable strings."""
    if value is None:
        return ""
    return str(value)


def _add_issue(issues: list[dict[str, Any]], issue_type: str, message: str, **extra: Any) -> None:
    """Append a structured workbook verification issue."""
    issue = {"type": issue_type, "message": message}
    issue.update(extra)
    issues.append(issue)
